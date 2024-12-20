import pywhatkit
import openpyxl
from datetime import datetime
import sys
import os
import time
import requests
from tkinter import *
import re
  
# import filedialog module
from tkinter import filedialog
from tkinter import messagebox

__author__ = "Nicholas Rademaker"
__copyright__ = None
__credits__ = [""]
__license__ = None
__version__ = "1.0.0"
__maintainer__ = "Nicholas Rademaker"
__email__ = "nicholasrademaker@hotmail.com"
__status__ = "Development"

class Simple():
    def __init__(self):
        self.API = "https://api.openweathermap.org/data/2.5/weather?q={city_name}&appid={api_key}&units=metric"
        self.key = "315323c54bacd4315a4e8c5be0845973"

        
    def comp(self, excel):
        wb = openpyxl.load_workbook(excel, data_only=True)
        sh = wb.active
        ws = sh
        
        fnames = []
        lnames = []
        phones = []
        msgs = []
        
        for col in ws['B']:
            fnames.append(col.value)        
        for col in ws['C']:
            lnames.append(col.value)        
        for col in ws['D']:
            if re.match(r"^(0|\+?44)\s?(?:\d\s?){9,11}$", col.value):
                phones.append(col.value)
            else:
                messagebox.showerror("UVW Mass messanger", "Invalid number %s detected in list"%col.value)
                exit()
        for col in ws['E']:
            msgs.append(col.value)

        for x in range(len(phones)):
            pywhatkit.sendwhatmsg_instantly(phones[x], msgs[x].format(fname=fnames[x],lname=lnames[x]), 10, True, 10)
            time.sleep(1)

        messagebox.showinfo("UVW Mass messanger", "Mass delivery completed")

    def sender(self, city: str) -> dict:
        url = self.API.format(city_name=city, api_key=self.key)
        resp = requests.get(url)
        return resp.json()


    def start(self):
        print("Welcome to the interface")

        # Getting the current date and time
        now = datetime.now()

        # Format current hour and minute from current system time
        hour = now.strftime("%H")
        mins = now.strftime("%M")
        # May add verbose call
        #print("Current hour =", hour)
        #print("Current mins =", mins)

        hour = int(hour)
        
        mins = int(mins) + 1

        # Establishing delivery phone number
        #phone_number = "+447884553443" # must be a text type petros
        phone_number = "+447444712726" # must be a text type petros

        # Establishing sent message
        msg = "Test message for UVW"

        # Establishing Delay

        delay = 1

        # Delayed Sending protocol
        # In case of multiple messages keep window open until last message sent

        # pywhatkit.sendwhatmsg(phone_number, msg,hour, mins, delay, True, 0)


        # group message protocol template:
        # pywhatkit.sendwhatmsg_to_group("AB123CDEFGHijklmn", "hello", 12, 12, 30, True, 5)

        # Instant message protocol:
        # sendwhatmsg_instantly(phone_no: str, message: str, wait_time: int = 15, tab_close: bool = False, close_time: int = 3)
        # Backend code is written to throttle for 4 seconds or breaking
        # 7 seconds seems to be the sweet spot

        pywhatkit.sendwhatmsg_instantly(phone_number, msg, 20, True, 20)
        # pywhatkit.core.close_tab(wait_time=5)
        print("Message Successfully sent")

if __name__ == "__main__":
    #print('Arguements: %s'%sys.argv)
    #os.system('cls')
    messagebox.showinfo("UVW Mass messanger", "Hello welcome to the UVW mass messanger please remember to login to your whatsapp web account before continuing")
    app = Simple()
    if len(sys.argv) == 1:
        try:
            app.comp(filedialog.askopenfilename(initialdir = "./",
                                              title = "Select a File",
                                              filetypes = (("Excel files",
                                                            "*.xlsx*"),
                                                           ("all files",
                                                            "*.*"))))
        except openpyxl.utils.exceptions.InvalidFileException:
            messagebox.showerror("UVW Mass messanger", "You have selected an invalid filetype please restart and try again")
    else:
        app.comp(sys.argv[-1])



